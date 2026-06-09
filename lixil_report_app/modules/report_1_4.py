"""
Report 1.4 - All Agents States Month
Columns: Agent, Not Ready(Td), Ready(Td), Reserved(Td), Talking(Td), Work(Td), Total Login(Td)
Excel export: Table + Total row + Clustered bar chart + one pie chart per agent
"""
import io
import pandas as pd
from .utils import read_cuic_xls, timedelta_to_hms, timedelta_to_hours, df_to_records


def process(file_data: bytes, filename: str = '') -> dict:
    df = read_cuic_xls(file_data)

    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if cl == 'agent':
            col_map[c] = 'Agent'
        elif 'not ready' in cl:
            col_map[c] = 'Not Ready'
        elif 'ready' in cl and 'not' not in cl and 'total' not in cl:
            col_map[c] = 'Ready'
        elif 'reserved' in cl:
            col_map[c] = 'Reserved'
        elif 'talking' in cl or 'talk' in cl:
            col_map[c] = 'Talking'
        elif 'work' in cl and 'total' not in cl:
            col_map[c] = 'Work'
        elif 'total' in cl or 'login' in cl:
            col_map[c] = 'Total Login'
    df = df.rename(columns=col_map)

    if 'Agent' in df.columns:
        df = df[df['Agent'].notna() & (df['Agent'].astype(str).str.strip() != '')].reset_index(drop=True)

    dur_cols = ['Not Ready', 'Ready', 'Reserved', 'Talking', 'Work', 'Total Login']
    dur_h = {}
    for c in dur_cols:
        if c in df.columns:
            dur_h[c] = df[c].apply(timedelta_to_hours).tolist()
            df[c] = df[c].apply(timedelta_to_hms)

    agents = df['Agent'].tolist() if 'Agent' in df.columns else [str(i) for i in range(len(df))]
    out = [c for c in ['Agent'] + dur_cols if c in df.columns]

    STATE_DISPLAY = ['Not Ready', 'Ready', 'Reserved', 'Talking', 'Work']
    pie_data = []
    for ai, agent in enumerate(agents):
        slices_raw = []
        for label in STATE_DISPLAY:
            hours_list = dur_h.get(label, [])
            try:
                val = round(float(hours_list[ai]), 2)
            except Exception:
                val = 0.0
            slices_raw.append({'label': label, 'value': val})
        total = sum(s['value'] for s in slices_raw) or 1.0
        slices = [{'label': s['label'], 'value': round(s['value'] / total * 100, 2)}
                  for s in slices_raw]
        pie_data.append({'agent': agent, 'slices': slices})

    # Compute column totals for web table display
    web_totals = {}
    for c in dur_cols:
        if c in df.columns:
            total_sec = sum(_hms_to_sec(v) for v in df[c])
            web_totals[c] = _sec_to_hms(total_sec)

    return {
        'report': '1_4',
        'title': '1.4 All Agents States Month - สถานะเจ้าหน้าที่ทุกคนรายเดือน',
        'labels': agents,
        'agents': agents,
        'state_cols': STATE_DISPLAY,
        'totals': web_totals,
        'datasets': {
            'not_ready': dur_h.get('Not Ready', []),
            'ready':     dur_h.get('Ready', []),
            'reserved':  dur_h.get('Reserved', []),
            'talking':   dur_h.get('Talking', []),
            'work':      dur_h.get('Work', []),
            'Not Ready': dur_h.get('Not Ready', []),
            'Ready':     dur_h.get('Ready', []),
            'Reserved':  dur_h.get('Reserved', []),
            'Talking':   dur_h.get('Talking', []),
            'Work':      dur_h.get('Work', []),
        },
        'pie_data': pie_data,
        'table': df_to_records(df[out]), 'columns': out,
    }


def _hms_to_sec(hms) -> int:
    try:
        parts = str(hms).strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except Exception:
        pass
    return 0


def _sec_to_hms(sec: int) -> str:
    sec = int(sec)
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f'{h:02d}:{m:02d}:{s:02d}'


def _fix_chart_axes_xml(xlsx_bytes: bytes) -> bytes:
    import zipfile as _zf
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml = raw.decode('utf-8')
                cs = xml.find('<catAx>')
                ce = xml.find('</catAx>') + len('</catAx>')
                if cs >= 0:
                    cat = xml[cs:ce]
                    cat = cat.replace(
                        '<axPos val="l"/>',
                        '<axPos val="b"/><delete val="0"/><crosses val="autoZero"/>'
                    )
                    xml = xml[:cs] + cat + xml[ce:]
                vs = xml.find('<valAx>')
                ve = xml.find('</valAx>') + len('</valAx>')
                if vs >= 0:
                    val = xml[vs:ve]
                    if '<delete' not in val:
                        val = val.replace(
                            '<axPos val="l"/>',
                            '<axPos val="l"/><delete val="0"/><crosses val="autoZero"/>'
                        )
                    xml = xml[:vs] + val + xml[ve:]
                raw = xml.encode('utf-8')
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.series import SeriesLabel

    data     = payload.get('data', payload)
    records  = data.get('table', [])
    cols     = data.get('columns', list(records[0].keys()) if records else [])
    datasets = data.get('datasets', {})
    title    = data.get('title', '1.4 All Agents States Month')

    ORANGE  = 'F15A22'
    NAVY    = '0D1B35'
    LGRAY   = 'F2F2F2'
    WHITE   = 'FFFFFF'
    LBLUE   = 'BDD7EE'
    BLUE    = '1F4E79'

    STATE_COLORS = ['FF0000', 'FF99FF', 'FFC000', '70AD47', '4472C4']
    STATE_KEYS   = ['not_ready', 'ready', 'reserved', 'talking', 'work']
    STATE_NAMES  = ['Not Ready', 'Ready', 'Reserved', 'Talking', 'Work']
    DUR_COLS     = {'Not Ready', 'Ready', 'Reserved', 'Talking', 'Work', 'Total Login'}

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'All Agents Month'

    agents   = [str(r.get('Agent', '')) for r in records]
    n_agents = len(agents)
    n_states = len(STATE_KEYS)

    # ── Build Total row ───────────────────────────────────────────────────
    total_row = {c: '' for c in cols}
    if cols:
        total_row[cols[0]] = 'Total'
    for col in cols:
        if col in DUR_COLS:
            total_sec = sum(_hms_to_sec(r.get(col, '00:00:00')) for r in records)
            total_row[col] = _sec_to_hms(total_sec)

    all_rows = list(records) + [total_row]

    # ── Row 1: Title ──────────────────────────────────────────────────────
    n_cols = len(cols)
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color=NAVY)
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # ── Row 2: Header ─────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(2, ci, col)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[2].height = 30

    # ── Rows 3+: Data + Total ─────────────────────────────────────────────
    for ri, rec in enumerate(all_rows, start=3):
        is_total = (ri == 2 + len(all_rows))
        for ci, col in enumerate(cols, 1):
            val = rec.get(col, '')
            cell = ws.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_total:
                cell.font = Font(bold=True, color=BLUE)
                cell.fill = PatternFill('solid', fgColor=LBLUE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LGRAY)

    # ── Column widths ─────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        w = 22 if col == 'Agent' else 14
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Hidden sheet for chart data ───────────────────────────────────────
    wd = wb.create_sheet('_data')
    wd.sheet_state = 'hidden'

    # Hours block: row 1 = agent headers, rows 2..n_states+1 = hours per state
    wd.cell(1, 1, 'Agent')
    for ai, agent in enumerate(agents):
        wd.cell(1, 2 + ai, agent)

    for si, key in enumerate(STATE_KEYS):
        row = 2 + si
        wd.cell(row, 1, STATE_NAMES[si])
        hours_list = datasets.get(key, [])
        for ai, h in enumerate(hours_list):
            try:
                wd.cell(row, 2 + ai, round(float(h), 2))
            except Exception:
                wd.cell(row, 2 + ai, 0.0)

    # Percentage block: row P_START = agent headers, rows P_START+1..+n_states = % per state
    P_START = n_states + 3
    wd.cell(P_START, 1, 'State')
    for ai, agent in enumerate(agents):
        wd.cell(P_START, 2 + ai, agent)

    for ai in range(n_agents):
        totals = []
        for key in STATE_KEYS:
            try:
                totals.append(float(datasets.get(key, [])[ai]))
            except Exception:
                totals.append(0.0)
        grand = sum(totals) or 1.0
        for si, t in enumerate(totals):
            row = P_START + 1 + si
            wd.cell(row, 1, STATE_NAMES[si])
            wd.cell(row, 2 + ai, round(t / grand * 100, 2))

    # ── Clustered Bar Chart ───────────────────────────────────────────────
    table_end_row  = 2 + len(all_rows)
    bar_anchor_row = table_end_row + 3

    if n_agents > 0:
        bar = BarChart()
        bar.type      = 'col'
        bar.grouping  = 'clustered'
        bar.title     = title
        bar.y_axis.title = 'Hours'
        bar.x_axis.title = 'Agent'
        bar.width  = max(20, n_agents * 2.5)
        bar.height = 14

        for si in range(n_states):
            ref = Reference(wd, min_col=2, max_col=1 + n_agents,
                            min_row=2 + si, max_row=2 + si)
            bar.add_data(ref, titles_from_data=False, from_rows=True)
            bar.series[-1].tx = SeriesLabel(v=STATE_NAMES[si])
            color = STATE_COLORS[si]
            bar.series[-1].graphicalProperties.solidFill = color
            bar.series[-1].graphicalProperties.line.solidFill = color

        bar.set_categories(Reference(wd, min_col=2, max_col=1 + n_agents,
                                     min_row=1, max_row=1))
        ws.add_chart(bar, f'A{bar_anchor_row}')

        # ── Pie Charts (2 per row) ─────────────────────────────────────────
        bar_height_rows = 22
        pie_start_row   = bar_anchor_row + bar_height_rows
        PIE_W, PIE_H, COLS_PER_PIE = 12, 12, 6

        for ai, agent in enumerate(agents):
            pie = PieChart()
            pie.title  = agent
            pie.width  = PIE_W
            pie.height = PIE_H

            labels_ref = Reference(wd, min_col=1,
                                   min_row=P_START + 1,
                                   max_row=P_START + n_states)
            data_ref   = Reference(wd, min_col=2 + ai,
                                   min_row=P_START + 1,
                                   max_row=P_START + n_states)
            pie.add_data(data_ref, titles_from_data=False)
            pie.set_categories(labels_ref)

            col_pos    = (ai % 2) * COLS_PER_PIE
            row_pos    = pie_start_row + (ai // 2) * 18
            col_letter = get_column_letter(col_pos + 1)
            ws.add_chart(pie, f'{col_letter}{row_pos}')

    out = io.BytesIO()
    wb.save(out)
    return _fix_chart_axes_xml(out.getvalue())
