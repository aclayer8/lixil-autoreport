"""
Report 1.3 - Agent States Day (individual agent, daily)
Columns: Date, Not Ready Duration(Td), %Not Ready, Ready Duration(Td), %Ready,
         Reserved Duration(Td), %Reserved, Talk Duration(Td), %Talking,
         Work Duration(Td), %Work, Day Total(Td)
"""
import io
import pandas as pd
from .utils import read_cuic_xls, safe_float, timedelta_to_hms, df_to_records


def process(file_data: bytes, filename: str = '') -> dict:
    df = read_cuic_xls(file_data)

    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if cl == 'date':
            col_map[c] = 'Date'
        elif 'not ready duration' in cl:
            col_map[c] = 'Not Ready Dur'
        elif '%not ready' in cl:
            col_map[c] = '%Not Ready'
        elif 'ready duration' in cl and 'not' not in cl:
            col_map[c] = 'Ready Dur'
        elif '%ready' in cl and 'not' not in cl:
            col_map[c] = '%Ready'
        elif 'reserved duration' in cl:
            col_map[c] = 'Reserved Dur'
        elif '%reserved' in cl:
            col_map[c] = '%Reserved'
        elif 'talk duration' in cl:
            col_map[c] = 'Talk Dur'
        elif '%talking' in cl:
            col_map[c] = '%Talking'
        elif 'work duration' in cl:
            col_map[c] = 'Work Dur'
        elif '%work' in cl:
            col_map[c] = '%Work'
        elif 'day total' in cl or 'total' in cl:
            col_map[c] = 'Day Total'
    df = df.rename(columns=col_map)

    if 'Date' in df.columns:
        df = df[df['Date'].notna() & (df['Date'].astype(str).str.strip() != '')].reset_index(drop=True)

    for c in ['Not Ready Dur', 'Ready Dur', 'Reserved Dur', 'Talk Dur', 'Work Dur', 'Day Total']:
        if c in df.columns:
            df[c] = df[c].apply(timedelta_to_hms)

    for c in ['%Not Ready', '%Ready', '%Reserved', '%Talking', '%Work']:
        if c in df.columns:
            df[c] = df[c].apply(safe_float)

    def fmt(v):
        try:
            return pd.to_datetime(str(v).strip()).strftime('%d/%m')
        except Exception:
            return str(v)[:10]

    labels = [fmt(v) for v in df['Date']] if 'Date' in df.columns else [str(i + 1) for i in range(len(df))]

    out = [c for c in ['Date', 'Not Ready Dur', '%Not Ready', 'Ready Dur', '%Ready',
                        'Reserved Dur', '%Reserved', 'Talk Dur', '%Talking',
                        'Work Dur', '%Work', 'Day Total'] if c in df.columns]

    def s(col):
        return df[col].tolist() if col in df.columns else []

    return {
        'report': '1_3',
        'title': '1.3 Agent States Day - สถานะเจ้าหน้าที่รายวัน',
        'labels': labels,
        'datasets': {
            'not_ready': s('%Not Ready'), 'ready': s('%Ready'),
            'reserved': s('%Reserved'), 'talking': s('%Talking'), 'work': s('%Work'),
        },
        'table': df_to_records(df[out]), 'columns': out,
    }


def _hms_to_sec(hms):
    """Convert HH:MM:SS string to total seconds."""
    try:
        parts = str(hms).strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except Exception:
        pass
    return 0


def _sec_to_hms(sec):
    """Convert total seconds to HH:MM:SS string."""
    sec = int(sec)
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f'{h:02d}:{m:02d}:{s:02d}'


def _coerce_val(val):
    if val is None or val == '':
        return val
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == '':
        return s
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return s


def _fix_chart_axes_xml(xlsx_bytes: bytes) -> bytes:
    import zipfile as _zf
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml = raw.decode('utf-8')
                cs = xml.find('<catAx>')
                ce = xml.find('</catAx>') + len('</catAx>')
                if cs >= 0:
                    cat = xml[cs:ce]
                    cat = cat.replace(
                        '<axPos val="l"/>',
                        '<axPos val="b"/><delete val="0"/><crosses val="autoZero"/>'
                    )
                    xml = xml[:cs] + cat + xml[ce:]
                vs = xml.find('<valAx>')
                ve = xml.find('</valAx>') + len('</valAx>')
                if vs >= 0:
                    val = xml[vs:ve]
                    if '<delete' not in val:
                        val = val.replace(
                            '<axPos val="l"/>',
                            '<axPos val="l"/><delete val="0"/><crosses val="autoZero"/>'
                        )
                    xml = xml[:vs] + val + xml[ve:]
                raw = xml.encode('utf-8')
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data    = payload.get('data', payload)
    records = data.get('table', [])
    cols    = data.get('columns', list(records[0].keys()) if records else [])
    title   = data.get('title', '1.3 Agent States Day')

    ORANGE      = 'F15A22'
    LIGHT_BLUE  = 'BDD7EE'
    NAVY        = '0D1B35'
    BLUE        = '1F4E79'
    LGRAY       = 'F2F2F2'
    WHITE       = 'FFFFFF'

    DUR_COLS = {'Not Ready Dur', 'Ready Dur', 'Reserved Dur', 'Talk Dur', 'Work Dur', 'Day Total'}
    PCT_COLS = {'%Not Ready', '%Ready', '%Reserved', '%Talking', '%Work'}

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Agent States Day'

    n_data = len(records)

    # Build Total row
    total_row = {c: '' for c in cols}
    if cols:
        total_row[cols[0]] = 'Total'

    for col in cols:
        if col in DUR_COLS:
            total_sec = sum(_hms_to_sec(r.get(col, '00:00:00')) for r in records)
            total_row[col] = _sec_to_hms(total_sec)
        elif col in PCT_COLS:
            vals = [_coerce_val(r.get(col, 0)) for r in records]
            numeric = [v for v in vals if isinstance(v, (int, float))]
            total_row[col] = round(sum(numeric), 2) if numeric else ''

    all_rows = list(records) + [total_row]

    # Row 1: Title
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color=NAVY)
    if len(cols) > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # Row 2: Header
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(2, ci, col)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[2].height = 30

    # Rows 3+: Data + Total
    for ri, rec in enumerate(all_rows, start=3):
        is_total = (ri == 2 + len(all_rows))
        for ci, col in enumerate(cols, 1):
            raw = rec.get(col)
            # Duration cols: keep as string; % cols: coerce to number
            val = raw if col in DUR_COLS else _coerce_val(raw)
            cell = ws.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_total:
                cell.font = Font(bold=True, color=BLUE)
                cell.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LGRAY)

    # Column widths
    for ci, col in enumerate(cols, 1):
        if col == 'Date':
            w = 14
        elif col in DUR_COLS:
            w = 14
        else:
            w = max(len(str(col)) + 2, 11)
        ws.column_dimensions[get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
