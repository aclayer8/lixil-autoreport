"""CSAT Report processor.

Replicates the delivered legacy CSAT page calculation:
- import survey sessions from Excel/CSV
- score 0 is counted separately
- detractors are 1-6, passives are 7-8, promoters are 9-10
- Total Calls is an editable value, defaulting to imported survey rows
"""
import io
import re
import pandas as pd
from .utils import safe_int

SESSION_COLUMNS = [
    'Node ID - Session ID - Sequence No',
    'Time',
    'Application Name',
    'Agent Name',
    'Agent Extension',
    'Calling Number',
    'Score',
]


def _norm(value) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())


ALIASES = {
    'Node ID - Session ID - Sequence No': {
        'nodeidsessionidsequenceno',
        'nodeidsessionidsequencenumber',
        'nodeidsessionsequenceno',
        'sessionidsequenceno',
        'sessionid',
    },
    'Time': {'time', 'datetime', 'date', 'calldatetime', 'starttime', 'endtime', 'timestamp'},
    'Application Name': {'applicationname', 'application', 'appname', 'applicationid'},
    'Agent Name': {'agentname', 'agent', 'agentfullname', 'resource'},
    'Agent Extension': {'agentextension', 'extension', 'agentext', 'ext', 'agentid', 'resourceid'},
    'Calling Number': {'callingnumber', 'callingno', 'callernumber', 'callerid', 'callingparty', 'ani', 'phone', 'phonenumber', 'number'},
    'Score': {'score', 'scores', 'point', 'points', 'rating', 'rate', 'csat', 'csatscore', 'surveyresult'},
}

THAI_ALIASES = {
    'Score': ('คะแนน', 'ความพึงพอใจ'),
    'Time': ('เวลา', 'วันที่'),
    'Agent Name': ('ชื่อเอเจนต์', 'เอเจนต์'),
    'Agent Extension': ('เบอร์เอเจนต์', 'ส่วนขยาย'),
    'Calling Number': ('เบอร์โทร', 'เบอร์ลูกค้า', 'หมายเลขโทร'),
}


def _read_excel_with_header(file_data: bytes, engine: str) -> pd.DataFrame:
    probe = pd.read_excel(io.BytesIO(file_data), engine=engine, header=None, dtype=object)
    header_idx = _detect_header_row(probe)
    if header_idx is None:
        return pd.read_excel(io.BytesIO(file_data), engine=engine, dtype=object)
    df = pd.read_excel(io.BytesIO(file_data), engine=engine, header=header_idx, dtype=object)
    return df


def _read_csv_with_header(file_data: bytes, encoding: str) -> pd.DataFrame:
    probe = pd.read_csv(io.BytesIO(file_data), encoding=encoding, header=None, dtype=object)
    header_idx = _detect_header_row(probe)
    if header_idx is None:
        return pd.read_csv(io.BytesIO(file_data), encoding=encoding, dtype=object)
    return pd.read_csv(io.BytesIO(file_data), encoding=encoding, header=header_idx, dtype=object)


def _read_table(file_data: bytes, filename: str) -> pd.DataFrame:
    lower = filename.lower()
    errors = []

    if lower.endswith('.csv'):
        for enc in ('utf-8-sig', 'utf-8', 'cp874', 'latin-1'):
            try:
                return _read_csv_with_header(file_data, enc)
            except Exception as exc:
                errors.append(f'csv/{enc}: {exc}')

    for engine in ('openpyxl', 'xlrd'):
        try:
            return _read_excel_with_header(file_data, engine)
        except Exception as exc:
            errors.append(f'{engine}: {exc}')

    raise ValueError('Cannot read CSAT file. Tried: ' + '; '.join(errors))


def _column_target(value):
    normalized = _norm(value)
    original = str(value or '').strip().lower()
    for target, aliases in ALIASES.items():
        if normalized in aliases:
            return target
    for target, terms in THAI_ALIASES.items():
        if any(term in original for term in terms):
            return target
    return None


def _detect_header_row(df: pd.DataFrame):
    best_idx = None
    best_score = 0
    for idx, row in df.head(40).iterrows():
        targets = {_column_target(v) for v in row.tolist()}
        targets.discard(None)
        score = len(targets)
        if 'Score' in targets:
            score += 2
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx if best_score >= 2 else None


def _map_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    used = set()
    for col in df.columns:
        target = _column_target(col)
        if target and target not in used:
            rename[col] = target
            used.add(target)
    return df.rename(columns=rename)


def _score(value) -> int:
    if value is None or pd.isna(value):
        return 0
    text = str(value).strip()
    if not text:
        return 0
    match = re.search(r'-?\d+', text)
    if not match:
        return 0
    return safe_int(match.group(0), 0)


def _looks_like_score_series(series: pd.Series) -> int:
    count = 0
    for value in series.dropna().head(100):
        score = _score(value)
        if 0 <= score <= 10 and str(value).strip() != '':
            count += 1
    return count


def _ensure_score_column(df: pd.DataFrame) -> pd.DataFrame:
    if 'Score' in df.columns:
        return df
    best_col = None
    best_count = 0
    for col in df.columns:
        count = _looks_like_score_series(df[col])
        if count > best_count:
            best_col = col
            best_count = count
    if best_col is not None and best_count > 0:
        return df.rename(columns={best_col: 'Score'})
    raise ValueError('Missing CSAT score column')


def _clean_cell(value):
    if value is None or pd.isna(value):
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _records(df: pd.DataFrame) -> pd.DataFrame:
    df = _map_columns(df)
    df = _ensure_score_column(df)
    df = df.dropna(how='all').reset_index(drop=True)

    for col in SESSION_COLUMNS:
        if col not in df.columns:
            df[col] = ''

    out = df[SESSION_COLUMNS].copy()
    out['Score'] = out['Score'].apply(_score)

    def is_data_row(row):
        node = str(row['Node ID - Session ID - Sequence No'] or '').strip().lower()
        if node.startswith(('generated on', 'filter ', 'interval ', 'report generated')):
            return False
        identity_fields = [
            'Node ID - Session ID - Sequence No', 'Time', 'Application Name',
            'Agent Name', 'Agent Extension', 'Calling Number'
        ]
        return any(str(row[col] or '').strip() for col in identity_fields)

    out = out[out.apply(is_data_row, axis=1)]
    return out.reset_index(drop=True)


def _df_to_records(df: pd.DataFrame) -> list:
    records = []
    for _, row in df.iterrows():
        item = {}
        for col in SESSION_COLUMNS:
            item[col] = _score(row[col]) if col == 'Score' else _clean_cell(row[col])
        records.append(item)
    return records


def calculate_metrics(records: list, total_calls=None) -> dict:
    scores = [_score(row.get('Score')) for row in records]
    total_survey = len(scores)
    total_calls = safe_int(total_calls, total_survey)
    point0 = sum(1 for s in scores if s == 0)
    point1_6 = sum(1 for s in scores if 1 <= s <= 6)
    point7_8 = sum(1 for s in scores if 7 <= s <= 8)
    point9_10 = sum(1 for s in scores if 9 <= s <= 10)
    answered = total_survey - point0
    promoters = (point9_10 / answered * 100) if answered else 0
    detractors = (point1_6 / answered * 100) if answered else 0
    response_rate = (answered / total_survey * 100) if total_survey else 0
    total_calls_survey_pct = (total_survey / total_calls * 100) if total_calls else 0
    total_call_response_rate = ((point1_6 + point7_8 + point9_10) / total_calls * 100) if total_calls else 0
    return {
        'totalCalls': total_calls,
        'totalCallsSurvey': total_survey,
        'point0': point0,
        'point1_6': point1_6,
        'point7_8': point7_8,
        'point9_10': point9_10,
        'totalCallsSurveyPercentage': round(total_calls_survey_pct, 2),
        'totalPercentageOfPromoters': round(promoters, 2),
        'totalPercentageOfDetractors': round(detractors, 2),
        'netPromoterScore': round(promoters - detractors, 2),
        'responseRate': round(response_rate, 2),
        'totalCallResponseRate': round(total_call_response_rate, 2),
    }


def process(file_data: bytes, filename: str = '') -> dict:
    raw = _read_table(file_data, filename)
    table = _records(raw)
    records = _df_to_records(table)
    metrics = calculate_metrics(records)
    return {
        'report': 'csat',
        'title': 'CSAT Report',
        'metrics': metrics,
        'table': records,
        'columns': SESSION_COLUMNS,
    }


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = payload.get('data', payload)
    records = data.get('table', [])
    metrics = data.get('metrics') or calculate_metrics(records)
    cols = data.get('columns', SESSION_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = 'CSAT Report'

    header_fill = PatternFill('solid', fgColor='0D1B35')
    orange_fill = PatternFill('solid', fgColor='F15A22')
    white_font = Font(color='FFFFFF', bold=True)
    border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))

    summary_cols = ['Total Calls', 'Total Calls Survey', 'Point 0', 'Point 1 - 6', 'Point 7 - 8', 'Point 9 - 10']
    summary_vals = [metrics['totalCalls'], metrics['totalCallsSurvey'], metrics['point0'],
                    metrics['point1_6'], metrics['point7_8'], metrics['point9_10']]
    calc_cols = ['Total Calls Survey (%)', 'Total % of Promoters', 'Total % of Detractors',
                 'Net Promoter Score', 'Response Rate', 'Total Call Response Rate']
    calc_vals = [metrics['totalCallsSurveyPercentage'], metrics['totalPercentageOfPromoters'],
                 metrics['totalPercentageOfDetractors'], metrics['netPromoterScore'],
                 metrics['responseRate'], metrics['totalCallResponseRate']]

    for c, name in enumerate(summary_cols, 9):
        cell = ws.cell(1, c, name)
        cell.fill = orange_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')
        ws.cell(2, c, summary_vals[c - 9])
    for c, name in enumerate(calc_cols, 9):
        cell = ws.cell(4, c, name)
        cell.fill = orange_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')
        ws.cell(5, c, calc_vals[c - 9])

    for c, name in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for r_idx, row in enumerate(records, 2):
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(r_idx, c_idx, row.get(col, ''))
            cell.border = border

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 42)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
