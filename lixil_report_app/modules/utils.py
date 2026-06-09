"""
Utility functions for reading CUIC XLS exports and building Excel responses.
CUIC exports XML Spreadsheet 2003 files with .xls extension.
Row 0 = blank, Row 1 = headers, Row 2+ = data, last N rows = notes.
"""
import io
import re
import xml.etree.ElementTree as ET
import pandas as pd
import numpy as np


def read_cuic_xls(file_data: bytes, header_row: int = 1, skip_footer: int = 2) -> pd.DataFrame:
    """
    Read a CUIC-exported file (.xls / .xlsx / .csv) into a DataFrame.
    Tries: XML Spreadsheet -> HTML -> xlrd -> openpyxl in order.
    """
    errors = []

    # --- Method 1: XML Spreadsheet 2003 ---
    try:
        df = _parse_xml_spreadsheet(file_data, header_row, skip_footer)
        if df is not None and len(df) > 0:
            return df
    except Exception as e:
        errors.append(f'xml: {e}')

    # --- Method 2: HTML table ---
    try:
        tables = pd.read_html(io.BytesIO(file_data), header=None)
        if tables:
            df = _apply_header(tables[0], header_row, skip_footer)
            if df is not None and len(df) > 0:
                return df
    except Exception as e:
        errors.append(f'html: {e}')

    # --- Method 3: Real binary XLS (BIFF) ---
    try:
        df = pd.read_excel(io.BytesIO(file_data), engine='xlrd',
                           header=header_row, skipfooter=skip_footer)
        df = _clean_df(df)
        if len(df) > 0:
            return df
    except Exception as e:
        errors.append(f'xlrd: {e}')

    # --- Method 4: XLSX ---
    try:
        df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl',
                           header=header_row, skipfooter=skip_footer)
        df = _clean_df(df)
        if len(df) > 0:
            return df
    except Exception as e:
        errors.append(f'openpyxl: {e}')

    raise ValueError("Cannot read file. Tried: " + "; ".join(errors))


def _parse_xml_spreadsheet(file_data: bytes, header_row: int, skip_footer: int):
    """Parse XML Spreadsheet 2003 (.xls exported by CUIC/web apps)."""
    text = file_data.decode('utf-8', errors='replace')

    if '<?xml' not in text and '<Workbook' not in text and '<Table' not in text:
        return None

    text = re.sub(r'\s+xmlns[^=\s]*="[^"]*"', '', text)
    text = re.sub(r'<(\w+):(\w+)', r'<\2', text)
    text = re.sub(r'</(\w+):(\w+)', r'</\2', text)
    text = re.sub(r'\s(\w+):(\w+)=', r' \2=', text)

    root = ET.fromstring(text)

    rows_data = []
    for row in root.iter('Row'):
        row_vals = []
        for cell in row.iter('Data'):
            row_vals.append(cell.text if cell.text is not None else '')
        if row_vals:
            rows_data.append(row_vals)

    if not rows_data:
        return None

    max_cols = max(len(r) for r in rows_data)
    rows_data = [r + [''] * (max_cols - len(r)) for r in rows_data]

    df_raw = pd.DataFrame(rows_data)
    return _apply_header(df_raw, header_row, skip_footer)


def _apply_header(df: pd.DataFrame, header_row: int, skip_footer: int):
    if df.shape[0] <= header_row:
        return None

    headers = [str(v).strip() for v in df.iloc[header_row]]
    data = df.iloc[header_row + 1:].copy()

    if skip_footer > 0 and len(data) > skip_footer:
        data = data.iloc[:-skip_footer]

    data.columns = headers
    return _clean_df(data)


def read_csv(file_data: bytes) -> pd.DataFrame:
    for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp874'):
        try:
            return pd.read_csv(io.BytesIO(file_data), encoding=enc)
        except Exception:
            continue
    raise ValueError("Cannot read CSV file")


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how='all').reset_index(drop=True)
    df = df.loc[:, df.notna().any()]
    df.columns = [str(c).strip() for c in df.columns]
    return df


def timedelta_to_hms(td, default='00:00:00') -> str:
    """Convert timedelta/Timedelta to HH:MM:SS string."""
    try:
        if td is None:
            return default
        try:
            is_na = bool(pd.isna(td))
            if is_na:
                return default
        except Exception:
            pass
        try:
            total_sec = int(pd.Timedelta(td).total_seconds())
        except Exception:
            total_sec = int(td.days * 86400 + td.seconds)
        h = total_sec // 3600
        m = (total_sec % 3600) // 60
        s = total_sec % 60
        return f'{h:02d}:{m:02d}:{s:02d}'
    except Exception:
        return default


def timedelta_to_hours(td, default=0.0) -> float:
    """Convert timedelta/Timedelta to decimal hours."""
    try:
        if td is None:
            return default
        try:
            is_na = bool(pd.isna(td))
            if is_na:
                return default
        except Exception:
            pass
        try:
            return round(pd.Timedelta(td).total_seconds() / 3600, 2)
        except Exception:
            return round((td.days * 86400 + td.seconds) / 3600, 2)
    except Exception:
        return default


def safe_float(val, default=0.0) -> float:
    try:
        if val is None:
            return default
        try:
            if bool(pd.isna(val)):
                return default
        except Exception:
            pass
        return float(str(val).replace('%', '').replace(',', '').strip())
    except Exception:
        return default


def safe_int(val, default=0) -> int:
    return int(safe_float(val, default))


def df_to_records(df: pd.DataFrame) -> list:
    """Convert DataFrame to JSON-safe list of dicts."""
    records = []
    for _, row in df.iterrows():
        rec = {}
        for col in df.columns:
            v = row[col]
            try:
                if pd.isna(v):
                    rec[col] = None
                    continue
            except Exception:
                pass
            if isinstance(v, (np.integer, int)):
                rec[col] = int(v)
            elif isinstance(v, (np.floating, float)):
                rec[col] = round(float(v), 4)
            else:
                rec[col] = str(v)
        records.append(rec)
    return records


def build_excel_response(df: pd.DataFrame, sheet_name: str = 'Report',
                          title: str = '', col_widths: dict = None) -> bytes:
    """Build a formatted .xlsx file from a DataFrame."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ORANGE = 'F15A22'
    NAVY   = '0D1B35'
    LGRAY  = 'F2F2F2'

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=ORANGE)
    title_font  = Font(bold=True, size=13, color=NAVY)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row_offset = 1
    if title:
        ws.cell(1, 1, title).font = title_font
        row_offset = 3

    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row_offset, ci, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin

    for ri, (_, row) in enumerate(df.iterrows(), start=row_offset + 1):
        fill = PatternFill('solid', fgColor=LGRAY) if ri % 2 == 0 else None
        for ci, col in enumerate(df.columns, start=1):
            cell = ws.cell(ri, ci, row[col])
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if fill:
                cell.fill = fill

    for ci, col in enumerate(df.columns, start=1):
        letter = get_column_letter(ci)
        w = col_widths.get(col, None) if col_widths else None
        ws.column_dimensions[letter].width = w if w else max(len(str(col)) + 4, 12)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
