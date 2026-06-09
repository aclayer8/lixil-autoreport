"""
Report 1.5 - Agent States YTD (HTML .xls format)
Columns: Month, %Not Ready, %Ready, %Reserved, %Talking, %Work
Excel export: Table + Line chart (% per state per month)
"""
import io
import pandas as pd
from .utils import safe_float, df_to_records


def _load(file_data: bytes) -> pd.DataFrame:
    try:
        tables = pd.read_html(io.BytesIO(file_data))
        if tables and len(tables[0]) > 0:
            df = tables[0]
            df.columns = [str(c).strip() for c in df.columns]
            return df.dropna(how='all').reset_index(drop=True)
    except Exception:
        pass
    try:
        df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl', header=1, skipfooter=2)
        df.columns = [str(c).strip() for c in df.columns]
        return df.dropna(how='all').reset_index(drop=True)
    except Exception:
        pass
    raise ValueError("Cannot read file 1.5")


def process(file_data: bytes, filename: str = '') -> dict:
    df = _load(file_data)

    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if 'month' in cl:
            col_map[c] = 'Month'
        elif 'not ready' in cl:
            col_map[c] = '%Not Ready'
        elif 'ready' in cl and 'not' not in cl:
            col_map[c] = '%Ready'
        elif 'reserved' in cl:
            col_map[c] = '%Reserved'
        elif 'talking' in cl or 'talk' in cl:
            col_map[c] = '%Talking'
        elif 'work' in cl:
            col_map[c] = '%Work'
    df = df.rename(columns=col_map)

    if 'Month' in df.columns:
        df = df[df['Month'].notna() & (df['Month'].astype(str).str.strip() != '')].reset_index(drop=True)

    pcts = ['%Not Ready', '%Ready', '%Reserved', '%Talking', '%Work']
    for c in pcts:
        if c in df.columns:
            df[c] = df[c].apply(safe_float)

    labels = df['Month'].astype(str).tolist() if 'Month' in df.columns else [str(i + 1) for i in range(len(df))]
    out = [c for c in ['Month'] + pcts if c in df.columns]

    def s(col):
        return df[col].tolist() if col in df.columns else []

    STATE_COLS = [c for c in pcts if c in df.columns]

    datasets = {
        # underscore keys (legacy)
        'not_ready': s('%Not Ready'), 'ready': s('%Ready'),
        'reserved':  s('%Reserved'),  'talking': s('%Talking'), 'work': s('%Work'),
    }
    # display-name keys for web chart (render_line_multi uses state_cols as lookup)
    for c in STATE_COLS:
        datasets[c] = s(c)

    # Average for total row
    totals = {}
    for c in STATE_COLS:
        vals = [v for v in s(c) if isinstance(v, (int, float))]
        totals[c] = round(sum(vals) / len(vals), 2) if vals else 0.0

    return {
        'report': '1_5',
        'title': '1.5 Agent States YTD - สถานะเจ้าหน้าที่ตั้งแต่ต้นปี',
        'labels': labels,
        'state_cols': STATE_COLS,
        'datasets': datasets,
        'totals': totals,
        'table': df_to_records(df[out]), 'columns': out,
    }


def _coerce_val(val):
    if val is None or val == '':
        return val
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == '':
        return s
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return s


def _fix_chart_axes_xml(xlsx_bytes: bytes) -> bytes:
    import zipfile as _zf
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml = raw.decode('utf-8')
                cs = xml.find('<catAx>')
                ce = xml.find('</catAx>') + len('</catAx>')
                if cs >= 0:
                    cat = xml[cs:ce]
                    cat = cat.replace(
                        '<axPos val="l"/>',
                        '<axPos val="b"/><delete val="0"/><crosses val="autoZero"/>'
                    )
                    xml = xml[:cs] + cat + xml[ce:]
                vs = xml.find('<valAx>')
                ve = xml.find('</valAx>') + len('</valAx>')
                if vs >= 0:
                    val = xml[vs:ve]
                    if '<delete' not in val:
                        val = val.replace(
                            '<axPos val="l"/>',
                            '<axPos val="l"/><delete val="0"/><crosses val="autoZero"/>'
                        )
                    xml = xml[:vs] + val + xml[ve:]
                raw = xml.encode('utf-8')
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    data    = payload.get('data', payload)
    records = data.get('table', [])
    cols    = data.get('columns', list(records[0].keys()) if records else [])
    title   = data.get('title', '1.5 Agent States YTD')

    ORANGE  = 'F15A22'
    NAVY    = '0D1B35'
    LGRAY   = 'F2F2F2'
    WHITE   = 'FFFFFF'

    # Line colors: Not Ready=red, Ready=green, Reserved=purple, Talking=orange, Work=blue
    STATE_COLS   = ['%Not Ready', '%Ready', '%Reserved', '%Talking', '%Work']
    LINE_COLORS  = ['FF0000', '70AD47', '9B59B6', 'F15A22', '4472C4']

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Agent States YTD'

    n_rows = len(records)
    n_cols = len(cols)

    # ── Build Average row ─────────────────────────────────────────────────
    avg_row = {c: '' for c in cols}
    if cols:
        avg_row[cols[0]] = 'Average'  # average of monthly %
    for col in cols:
        if col in STATE_COLS:
            vals = []
            for r in records:
                try:
                    v = float(r.get(col, 0) or 0)
                    vals.append(v)
                except Exception:
                    pass
            avg_row[col] = round(sum(vals) / len(vals), 2) if vals else 0.0

    all_rows = list(records) + [avg_row]

    # ── Row 1: Title ──────────────────────────────────────────────────────
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color=NAVY)
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # ── Row 2: Header ─────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(2, ci, col)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[2].height = 30

    # ── Rows 3+: Data + Average ───────────────────────────────────────────
    LBLUE = 'BDD7EE'
    BLUE  = '1F4E79'
    for ri, rec in enumerate(all_rows, start=3):
        is_avg = (ri == 2 + len(all_rows))
        for ci, col in enumerate(cols, 1):
            val = _coerce_val(rec.get(col, ''))
            cell = ws.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_avg:
                cell.font = Font(bold=True, color=BLUE)
                cell.fill = PatternFill('solid', fgColor=LBLUE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LGRAY)

    # ── Column widths ─────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        w = 14 if col == 'Month' else max(len(str(col)) + 4, 13)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Line Chart (data rows only, exclude Average row) ─────────────────
    if n_rows > 0:
        chart_series_cols = [i + 1 for i, c in enumerate(cols) if c in STATE_COLS]

        if chart_series_cols:
            chart = LineChart()
            chart.title    = title
            chart.y_axis.title = '%'
            chart.x_axis.title = 'Month'
            chart.width  = 26
            chart.height = 14
            chart.style  = 10

            for idx, col_idx in enumerate(chart_series_cols):
                col_name = cols[col_idx - 1]
                ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                                min_row=3, max_row=2 + n_rows)
                chart.add_data(ref, titles_from_data=False)
                s = chart.series[-1]
                s.tx = SeriesLabel(v=col_name)
                color = LINE_COLORS[idx % len(LINE_COLORS)]
                s.graphicalProperties.line.solidFill = color
                s.graphicalProperties.line.width = 20000   # 2pt
                s.smooth = True

            month_col_idx = (cols.index('Month') + 1) if 'Month' in cols else 1
            chart.set_categories(Reference(ws, min_col=month_col_idx,
                                           min_row=3, max_row=2 + n_rows))

            ws.add_chart(chart, f'A{n_rows + 5}')

    out = io.BytesIO()
    wb.save(out)
    return _fix_chart_axes_xml(out.getvalue())
