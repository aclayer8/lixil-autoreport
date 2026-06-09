"""
Report 1.2 - CSQ Hour
จำนวนสายเข้าเฉลี่ยในช่วงแต่ละชั่วโมง
Work hours: 08:00 AM - 16:00 (4 PM)
"""
import io
import pandas as pd
from .utils import read_cuic_xls, safe_float, df_to_records


def process(file_data: bytes, filename: str = '') -> dict:
    df = read_cuic_xls(file_data)

    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if 'time' in cl or 'date' in cl:
            col_map[c] = 'Time'
        elif 'presented' in cl:
            col_map[c] = 'Presented'
        elif 'handled' in cl:
            col_map[c] = 'Handled'
        elif 'abandoned' in cl:
            col_map[c] = 'Abandoned'
    df = df.rename(columns=col_map)

    for col in ['Presented', 'Handled', 'Abandoned']:
        if col not in df.columns:
            df[col] = 0.0
    df['Presented'] = df['Presented'].apply(safe_float)
    df['Handled']   = df['Handled'].apply(safe_float)
    df['Abandoned'] = df['Abandoned'].apply(safe_float)

    def to_raw_hour(v):
        try:
            return pd.Timestamp(v).hour
        except Exception:
            s = str(v)
            for p in s.split():
                if ':' in p and len(p) >= 5:
                    hh = p[:2]
                    if hh.isdigit():
                        return int(hh)
            return -1

    def work_hour_sort_key(h):
        if h < 0:
            return 99
        if 1 <= h <= 7:
            return h + 12
        return h

    def to_display_hour(h):
        if 1 <= h <= 7:
            return h + 12
        return h

    df['_rawHour'] = df['Time'].apply(to_raw_hour) if 'Time' in df.columns else -1
    df['HourKey']  = df['_rawHour'].apply(lambda h: f'{h:02d}:00:00' if h >= 0 else 'N/A')
    df['_sort']    = df['_rawHour'].apply(work_hour_sort_key)

    grouped = df.groupby('HourKey', sort=False).agg(
        Presented=('Presented', 'mean'),
        Handled=('Handled', 'mean'),
        Abandoned=('Abandoned', 'mean'),
        _rawHour=('_rawHour', 'first'),
        _sort=('_sort', 'first'),
    ).reset_index()
    grouped = grouped.sort_values('_sort').drop(columns='_sort').reset_index(drop=True)

    grouped['HourKey'] = grouped['_rawHour'].apply(
        lambda h: f'{to_display_hour(h):02d}:00:00' if h >= 0 else 'N/A'
    )
    grouped = grouped.drop(columns=['_rawHour'])
    grouped['Presented'] = grouped['Presented'].round(2)
    grouped['Handled']   = grouped['Handled'].round(2)
    grouped['Abandoned'] = grouped['Abandoned'].round(2)

    grand = {
        'HourKey':   'Grand Average',
        'Presented': round(grouped['Presented'].mean(), 2),
        'Handled':   round(grouped['Handled'].mean(), 2),
        'Abandoned': round(grouped['Abandoned'].mean(), 2),
    }
    grouped_with_grand = pd.concat([grouped, pd.DataFrame([grand])], ignore_index=True)

    grouped_with_grand['Time'] = grouped_with_grand['HourKey'].apply(
        lambda h: h if h == 'Grand Average' else f'{h} Average'
    )
    out_cols = ['Time', 'Presented', 'Handled', 'Abandoned']

    return {
        'report': '1_2',
        'title': '1.2 CSQ Hour - จำนวนสายเข้าเฉลี่ยในช่วงแต่ละชั่วโมง',
        'labels': grouped_with_grand['Time'].tolist(),
        'datasets': {
            'presented':  grouped_with_grand['Presented'].tolist(),
            'handled':    grouped_with_grand['Handled'].tolist(),
            'abandoned':  grouped_with_grand['Abandoned'].tolist(),
        },
        'table':   df_to_records(grouped_with_grand[out_cols]),
        'columns': out_cols,
    }


def _coerce_val(val):
    if val is None or val == '':
        return val
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == '':
        return s
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return s


def _fix_chart_axes_xml(xlsx_bytes: bytes) -> bytes:
    import zipfile as _zf
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml = raw.decode('utf-8')
                cs = xml.find('<catAx>')
                ce = xml.find('</catAx>') + len('</catAx>')
                if cs >= 0:
                    cat = xml[cs:ce]
                    cat = cat.replace(
                        '<axPos val="l"/>',
                        '<axPos val="b"/><delete val="0"/><crosses val="autoZero"/>'
                    )
                    xml = xml[:cs] + cat + xml[ce:]
                vs = xml.find('<valAx>')
                ve = xml.find('</valAx>') + len('</valAx>')
                if vs >= 0:
                    val = xml[vs:ve]
                    if '<delete' not in val:
                        val = val.replace(
                            '<axPos val="l"/>',
                            '<axPos val="l"/><delete val="0"/><crosses val="autoZero"/>'
                        )
                    xml = xml[:vs] + val + xml[ve:]
                raw = xml.encode('utf-8')
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    data    = payload.get('data', payload)
    records = data.get('table', [])
    cols    = data.get('columns', list(records[0].keys()) if records else [])
    title   = data.get('title', '1.2 CSQ Hour')

    ORANGE       = 'F15A22'
    LIGHT_ORANGE = 'FCE4D6'
    NAVY         = '0D1B35'
    LGRAY        = 'F2F2F2'
    WHITE        = 'FFFFFF'
    CHART_COLORS = ['4472C4', '548235', 'FFC000']

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'CSQ Hour'

    data_rows  = [r for r in records if str(r.get('Time', '')).strip() != 'Grand Average']
    grand_rows = [r for r in records if str(r.get('Time', '')).strip() == 'Grand Average']
    all_rows   = data_rows + grand_rows
    n_all      = len(all_rows)

    ws.cell(1, 1, title).font = Font(bold=True, size=13, color=NAVY)
    if len(cols) > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(2, ci, col)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[2].height = 25

    for ri, rec in enumerate(all_rows, start=3):
        is_grand = (str(rec.get('Time', '')).strip() == 'Grand Average')
        for ci, col in enumerate(cols, 1):
            val = _coerce_val(rec.get(col))
            cell = ws.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_grand:
                cell.font = Font(bold=True, color=NAVY)
                cell.fill = PatternFill('solid', fgColor=LIGHT_ORANGE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LGRAY)

    for ci, col in enumerate(cols, 1):
        width = 22 if col == 'Time' else max(len(str(col)) + 4, 14)
        ws.column_dimensions[get_column_letter(ci)].width = width

    if n_all > 0:
        chart_series_cols = ['Presented', 'Handled', 'Abandoned']
        chart_col_indices = [i + 1 for i, c in enumerate(cols) if c in chart_series_cols]

        if chart_col_indices:
            chart = BarChart()
            chart.type = 'col'
            chart.grouping = 'clustered'
            chart.title = title
            chart.y_axis.title = 'Number of Calls'
            chart.x_axis.title = 'Time'
            chart.width = 24
            chart.height = 14

            for col_idx in chart_col_indices:
                col_name = cols[col_idx - 1]
                ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                                min_row=3, max_row=2 + n_all)
                chart.add_data(ref, titles_from_data=False)
                chart.series[-1].tx = SeriesLabel(v=col_name)

            time_col_idx = (cols.index('Time') + 1) if 'Time' in cols else 1
            chart.set_categories(Reference(ws, min_col=time_col_idx,
                                           min_row=3, max_row=2 + n_all))

            for i, s in enumerate(chart.series):
                color = CHART_COLORS[i % len(CHART_COLORS)]
                s.graphicalProperties.solidFill = color
                s.graphicalProperties.line.solidFill = color

            ws.add_chart(chart, f'A{n_all + 5}')

    out = io.BytesIO()
    wb.save(out)
    return _fix_chart_axes_xml(out.getvalue())
