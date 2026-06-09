"""
Report 1.11 - Reserved/Talking Count Ratio per Agent
จำนวนครั้งที่สายเข้าจากคิว = COUNTIF(Agent State, "Reserved")
จำนวนครั้งที่รับสายจากคิว  = COUNTIF(Agent State, "Talking")
% รับสาย                    = (Talking × 100) / Reserved
"""
import io
import pandas as pd
from .utils import read_cuic_xls, df_to_records


def process(file_data: bytes, filename: str = '') -> dict:
    df = read_cuic_xls(file_data)

    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if 'agent name' in cl or cl == 'agent':
            col_map[c] = 'Agent Name'
        elif 'agent state' in cl or cl == 'state':
            col_map[c] = 'Agent State'
    df = df.rename(columns=col_map)

    if 'Agent Name' not in df.columns or 'Agent State' not in df.columns:
        raise ValueError("ไม่พบคอลัมน์ Agent Name หรือ Agent State")

    df = df[df['Agent Name'].notna() & (df['Agent Name'].astype(str).str.strip() != '')].copy()
    df['Agent State'] = df['Agent State'].astype(str).str.strip()

    rows = []
    for agent, grp in df.groupby('Agent Name', sort=True):
        reserved = int((grp['Agent State'] == 'Reserved').sum())
        talking  = int((grp['Agent State'] == 'Talking').sum())
        pct      = round((talking * 100) / reserved, 2) if reserved > 0 else 0.0
        rows.append({
            'Agent':                          str(agent),
            'จำนวนครั้งที่สายเข้าจากคิว':    reserved,
            'จำนวนครั้งที่รับสายจากคิว':     talking,
            '% รับสาย':                       pct,
        })

    cols = ['Agent', 'จำนวนครั้งที่สายเข้าจากคิว', 'จำนวนครั้งที่รับสายจากคิว', '% รับสาย']
    agents   = [r['Agent'] for r in rows]
    reserved = [r['จำนวนครั้งที่สายเข้าจากคิว'] for r in rows]
    talking  = [r['จำนวนครั้งที่รับสายจากคิว'] for r in rows]
    pct_list = [r['% รับสาย'] for r in rows]

    # Totals
    total_res  = sum(reserved)
    total_talk = sum(talking)
    total_pct  = round((total_talk * 100) / total_res, 2) if total_res > 0 else 0.0

    return {
        'report':  '1_11',
        'title':   '1.11 Reserved/Talking Ratio - อัตราการรับสายจากคิว',
        'labels':  agents,
        'datasets': {
            'reserved': reserved,
            'talking':  talking,
            'pct':      pct_list,
        },
        'totals': {
            'reserved': total_res,
            'talking':  total_talk,
            'pct':      total_pct,
        },
        'table':   rows,
        'columns': cols,
    }


def _fix_chart_axes_xml(xlsx_bytes: bytes) -> bytes:
    import zipfile as _zf
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml = raw.decode('utf-8')
                cs = xml.find('<catAx>')
                ce = xml.find('</catAx>') + len('</catAx>')
                if cs >= 0:
                    cat = xml[cs:ce]
                    cat = cat.replace('<axPos val="l"/>',
                                      '<axPos val="b"/><delete val="0"/><crosses val="autoZero"/>')
                    xml = xml[:cs] + cat + xml[ce:]
                vs = xml.find('<valAx>')
                ve = xml.find('</valAx>') + len('</valAx>')
                if vs >= 0:
                    val = xml[vs:ve]
                    if '<delete' not in val:
                        val = val.replace('<axPos val="l"/>',
                                          '<axPos val="l"/><delete val="0"/><crosses val="autoZero"/>')
                    xml = xml[:vs] + val + xml[ve:]
                raw = xml.encode('utf-8')
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    data    = payload.get('data', payload)
    records = data.get('table', [])
    cols    = data.get('columns', list(records[0].keys()) if records else [])
    totals  = data.get('totals', {})
    title   = data.get('title', '1.11 Reserved/Talking Ratio')

    ORANGE  = 'F15A22'
    NAVY    = '0D1B35'
    LGRAY   = 'F2F2F2'
    WHITE   = 'FFFFFF'
    LBLUE   = 'BDD7EE'
    BLUE    = '1F4E79'

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reserved Talking Ratio'

    n_data = len(records)
    n_cols = len(cols)

    # ── Total row ────────────────────────────────────────────────────────
    total_row = {c: '' for c in cols}
    total_row['Agent'] = 'Total'
    total_row['จำนวนครั้งที่สายเข้าจากคิว'] = totals.get('reserved', '')
    total_row['จำนวนครั้งที่รับสายจากคิว']  = totals.get('talking', '')
    total_row['% รับสาย']                    = totals.get('pct', '')

    all_rows = list(records) + [total_row]

    # ── Row 1: Title ─────────────────────────────────────────────────────
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color=NAVY)
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # ── Row 2: Header ────────────────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(2, ci, col)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[2].height = 35

    # ── Rows 3+: Data + Total ────────────────────────────────────────────
    for ri, rec in enumerate(all_rows, start=3):
        is_total = (ri == 2 + len(all_rows))
        for ci, col in enumerate(cols, 1):
            val = rec.get(col, '')
            cell = ws.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_total:
                cell.font = Font(bold=True, color=BLUE)
                cell.fill = PatternFill('solid', fgColor=LBLUE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LGRAY)

    # ── Column widths ────────────────────────────────────────────────────
    col_widths = {
        'Agent': 22,
        'จำนวนครั้งที่สายเข้าจากคิว': 22,
        'จำนวนครั้งที่รับสายจากคิว':  22,
        '% รับสาย': 14,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 16)

    # ── Clustered Bar Chart (Reserved vs Talking counts) ─────────────────
    if n_data > 0:
        agent_col_idx   = cols.index('Agent') + 1 if 'Agent' in cols else 1
        res_col_idx     = cols.index('จำนวนครั้งที่สายเข้าจากคิว') + 1
        talk_col_idx    = cols.index('จำนวนครั้งที่รับสายจากคิว') + 1

        chart = BarChart()
        chart.type      = 'col'
        chart.grouping  = 'clustered'
        chart.title     = title
        chart.y_axis.title = 'จำนวนครั้ง'
        chart.x_axis.title = 'Agent'
        chart.width  = max(20, n_data * 2.5)
        chart.height = 14

        for col_idx, label, color in [
            (res_col_idx,  'สายเข้าจากคิว (Reserved)', '9B59B6'),
            (talk_col_idx, 'รับสาย (Talking)',          'F15A22'),
        ]:
            ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                            min_row=3, max_row=2 + n_data)
            chart.add_data(ref, titles_from_data=False)
            chart.series[-1].tx = SeriesLabel(v=label)
            chart.series[-1].graphicalProperties.solidFill = color
            chart.series[-1].graphicalProperties.line.solidFill = color

        chart.set_categories(Reference(ws, min_col=agent_col_idx,
                                       min_row=3, max_row=2 + n_data))
        ws.add_chart(chart, f'A{n_data + 5}')

    out = io.BytesIO()
    wb.save(out)
    return _fix_chart_axes_xml(out.getvalue())
