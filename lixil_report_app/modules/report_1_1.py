"""
Report 1.1 - CSQ Day
จำนวนสายที่เข้ามาในคิวของคอนแทคเซ็นเตอร์ (รายวัน)
File: real .xlsx with header NOT at row 1 — find it dynamically.
Columns: Time, Presented, Handled, %Handled, Abandoned, %Abandoned
"""
import io
import pandas as pd
import openpyxl
from .utils import safe_float, safe_int, df_to_records


def _load(file_data: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]

    header_idx = None
    for i, row in enumerate(all_rows):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if any('presented' in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("ไม่พบแถว header ในไฟล์")

    headers = [str(v).strip() if v is not None else '' for v in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    # Strip footer rows: empty rows, CUIC metadata ("Generated on..."),
    # Thai summary rows ("รวม"), and blank-date subtotal rows — from the bottom up.
    while data_rows:
        last = data_rows[-1]
        if all(v is None or str(v).strip() == '' for v in last):
            data_rows.pop()
            continue
        first_val = str(last[0]).strip() if last[0] is not None else ''
        if (first_val.lower().startswith('generated') or
                'รวม' in first_val or
                first_val == ''):
            data_rows.pop()
            continue
        break

    df = pd.DataFrame(data_rows, columns=headers)
    return df.dropna(how='all').reset_index(drop=True)


def process(file_data: bytes, filename: str = '') -> dict:
    df = _load(file_data)

    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if cl in ('time', 'date'):
            col_map[c] = 'Time'
        elif '%handled' in cl or 'handled%' in cl or 'handled %' in cl:
            col_map[c] = 'Handled%'
        elif 'presented' in cl:
            col_map[c] = 'Presented'
        elif 'handled' in cl:
            col_map[c] = 'Handled'
        elif '%abandoned' in cl or 'abandoned%' in cl or 'abandoned %' in cl:
            col_map[c] = 'Abandoned%'
        elif 'abandoned' in cl:
            col_map[c] = 'Abandoned'
    df = df.rename(columns=col_map)

    for col in ['Presented', 'Handled', 'Abandoned']:
        if col not in df.columns:
            df[col] = 0
    df['Presented'] = df['Presented'].apply(safe_int)
    df['Handled']   = df['Handled'].apply(safe_int)
    df['Abandoned'] = df['Abandoned'].apply(safe_int)
    df['Handled%']  = df.apply(
        lambda r: round(r['Handled'] / r['Presented'] * 100, 2) if r['Presented'] > 0 else 0.0, axis=1)
    df['Abandoned%'] = df.apply(
        lambda r: round(r['Abandoned'] / r['Presented'] * 100, 2) if r['Presented'] > 0 else 0.0, axis=1)

    labels = []
    if 'Time' in df.columns:
        for v in df['Time']:
            s = str(v).strip()
            try:
                labels.append(pd.to_datetime(s, dayfirst=True).strftime('%d/%m'))
            except Exception:
                labels.append(s[:10] if len(s) >= 10 else s)
    else:
        labels = [str(i + 1) for i in range(len(df))]

    total_p = int(df['Presented'].sum())
    total_h = int(df['Handled'].sum())
    total_a = int(df['Abandoned'].sum())

    out_cols = [c for c in ['Time', 'Presented', 'Handled', 'Handled%', 'Abandoned', 'Abandoned%']
                if c in df.columns]

    return {
        'report': '1_1',
        'title': '1.1 CSQ Day - จำนวนสายที่เข้ามาในคิว',
        'labels': labels,
        'datasets': {
            'presented':    df['Presented'].tolist(),
            'handled':      df['Handled'].tolist(),
            'abandoned':    df['Abandoned'].tolist(),
            'handled_pct':  df['Handled%'].tolist(),
            'abandoned_pct': df['Abandoned%'].tolist(),
        },
        'totals': {
            'presented': total_p, 'handled': total_h, 'abandoned': total_a,
            'handled_pct':   round(total_h / total_p * 100, 2) if total_p else 0.0,
            'abandoned_pct': round(total_a / total_p * 100, 2) if total_p else 0.0,
        },
        'table':   df_to_records(df[out_cols]),
        'columns': out_cols,
    }


def _coerce_val(val):
    """Convert string-numbers to int/float so openpyxl writes them as numbers, not text."""
    if val is None or val == '':
        return val
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == '':
        return s
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return s


def _fix_chart_axes_xml(xlsx_bytes: bytes) -> bytes:
    """Fix openpyxl chart XML for Excel: catAx needs axPos=b, plus delete+crosses."""
    import zipfile as _zf
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml = raw.decode('utf-8')
                cs = xml.find('<catAx>')
                ce = xml.find('</catAx>') + len('</catAx>')
                if cs >= 0:
                    cat = xml[cs:ce]
                    cat = cat.replace(
                        '<axPos val="l"/>',
                        '<axPos val="b"/><delete val="0"/><crosses val="autoZero"/>'
                    )
                    xml = xml[:cs] + cat + xml[ce:]
                vs = xml.find('<valAx>')
                ve = xml.find('</valAx>') + len('</valAx>')
                if vs >= 0:
                    val = xml[vs:ve]
                    if '<delete' not in val:
                        val = val.replace(
                            '<axPos val="l"/>',
                            '<axPos val="l"/><delete val="0"/><crosses val="autoZero"/>'
                        )
                    xml = xml[:vs] + val + xml[ve:]
                raw = xml.encode('utf-8')
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    data    = payload.get('data', payload)
    records = data.get('table', [])
    cols    = data.get('columns', list(records[0].keys()) if records else [])
    totals  = data.get('totals', {})
    title   = data.get('title', '1.1 CSQ Day')

    ORANGE       = 'F15A22'
    LIGHT_ORANGE = 'FCE4D6'
    NAVY         = '0D1B35'
    LGRAY        = 'F2F2F2'
    WHITE        = 'FFFFFF'
    # Presented=blue, Handled=orange, Abandoned=yellow (matching 1.1.8 standard colors)
    CHART_COLORS = ['4472C4', 'ED7D31', 'FFC000']

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'CSQ Day'

    n_data = len(records)   # data rows only (before adding Total row)

    # Build total row
    total_row = {c: '' for c in cols}
    if cols:
        total_row[cols[0]] = 'Total'
    for k, tk in [('Presented', 'presented'), ('Handled', 'handled'),
                  ('Abandoned', 'abandoned'), ('Handled%', 'handled_pct'),
                  ('Abandoned%', 'abandoned_pct')]:
        if k in total_row:
            total_row[k] = totals.get(tk, '')

    all_rows = list(records) + [total_row]

    # Row 1: Title (merged)
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color=NAVY)
    if len(cols) > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # Row 2: Header
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(2, ci, col)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[2].height = 25

    # Rows 3+: Data + Total — coerce values to proper types before writing
    for ri, rec in enumerate(all_rows, start=3):
        is_total = (ri == 2 + len(all_rows))
        for ci, col in enumerate(cols, 1):
            val = _coerce_val(rec.get(col))
            cell = ws.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_total:
                cell.font = Font(bold=True, color=NAVY)
                cell.fill = PatternFill('solid', fgColor=LIGHT_ORANGE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LGRAY)

    # Column widths
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col)) + 4, 14)

    # Chart (1.1.4): Clustered Column — Presented, Handled, Abandoned only
    if n_data > 0:
        chart_series_cols = ['Presented', 'Handled', 'Abandoned']
        chart_col_indices = [i + 1 for i, c in enumerate(cols) if c in chart_series_cols]

        if chart_col_indices:
            chart = BarChart()
            chart.type = 'col'
            chart.grouping = 'clustered'
            chart.title = title
            chart.y_axis.title = 'Number of Calls'
            chart.x_axis.title = 'Date'
            chart.width = 24
            chart.height = 14

            # Add each series — data rows only, exclude Total row
            for col_idx in chart_col_indices:
                col_name = cols[col_idx - 1]
                ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                                min_row=3, max_row=2 + n_data)
                chart.add_data(ref, titles_from_data=False)
                chart.series[-1].tx = SeriesLabel(v=col_name)

            # Date labels from Time column
            time_col_idx = (cols.index('Time') + 1) if 'Time' in cols else 1
            chart.set_categories(Reference(ws, min_col=time_col_idx,
                                           min_row=3, max_row=2 + n_data))

            # Apply colors
            for i, s in enumerate(chart.series):
                color = CHART_COLORS[i % len(CHART_COLORS)]
                s.graphicalProperties.solidFill = color
                s.graphicalProperties.line.solidFill = color

            ws.add_chart(chart, f'A{n_data + 5}')

    out = io.BytesIO()
    wb.save(out)
    return _fix_chart_axes_xml(out.getvalue())
