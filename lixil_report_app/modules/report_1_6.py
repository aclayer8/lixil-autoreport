"""
Report 1.6 - Not Ready Count (ตารางรายวัน + YTD)
"""
import re
import io
import pandas as pd
from .utils import read_cuic_xls, df_to_records

REASON_TO_COL = {
    'service customer': 'Service Customers', 'wrap up time expiry': 'Service Customers',
    'offhook': 'Service Customers', 'outbound': 'Service Customers',
    'lunch': 'Lunch', 'meeting': 'Meeting', 'break': 'Break',
    'non acd busy': 'Inbound From Agent',
    '600': 'Service Customers', '32758': 'Service Customers',
    '32762': 'Service Customers', '32760': 'Service Customers',
    '601': 'Lunch', '602': 'Meeting', '603': 'Break', '32761': 'Inbound From Agent',
}
DISPLAY_COLS = ['Service Customers', 'Lunch', 'Meeting', 'Break', 'Inbound From Agent']
PCT_COLS = ['Service Customers', 'Lunch', 'Meeting', 'Break']


def _extract_agent_name(filename):
    try:
        name = re.sub(r'\.(?:xlsx?|csv)$', '', filename, flags=re.IGNORECASE)
        m = re.match(r'1\.6\s+(.+?)-', name, re.IGNORECASE)
        if m:
            return m.group(1).strip()
        parts = name.split()
        if len(parts) >= 2:
            return parts[1]
    except Exception:
        pass
    return 'Agent'


def _fmt_date(dt):
    try:
        return f"{dt.month}/{dt.day}/{str(dt.year)[-2:]}"
    except Exception:
        return str(dt)


def _map_reason(raw):
    return REASON_TO_COL.get(str(raw).strip().lower(), None)


def process(file_data, filename=''):
    df = read_cuic_xls(file_data)
    col_map = {}
    for c in df.columns:
        cl = c.lower().strip()
        if 'reason' in cl:
            col_map[c] = 'reason_code'
        elif 'state' in cl:
            col_map[c] = 'agent_state'
        elif 'transition' in cl or 'time' in cl:
            col_map[c] = 'transition_time'
    df = df.rename(columns=col_map)
    if 'agent_state' in df.columns:
        nr = df[df['agent_state'].astype(str).str.contains('Not Ready', na=False)].copy()
    else:
        nr = df.copy()
    if 'reason_code' not in nr.columns or 'transition_time' not in nr.columns:
        return _empty_result()
    nr['display_col'] = nr['reason_code'].apply(_map_reason)
    nr = nr[nr['display_col'].notna()].copy()
    if len(nr) == 0:
        return _empty_result()
    nr['date'] = pd.to_datetime(nr['transition_time'].astype(str).str.strip(), format='%m/%d/%y', errors='coerce')
    nr = nr[nr['date'].notna()].copy()
    if len(nr) == 0:
        return _empty_result()
    nr['month'] = nr['date'].dt.to_period('M')
    agent_name = _extract_agent_name(filename)
    all_months = sorted(nr['month'].unique())
    report_month = all_months[-1]
    month_df = nr[nr['month'] == report_month].copy()
    month_df['date_str'] = month_df['date'].apply(_fmt_date)
    pivot = month_df.groupby(['date_str', 'display_col']).size().unstack(fill_value=0)
    for col in DISPLAY_COLS:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot[DISPLAY_COLS].reset_index().rename(columns={'date_str': 'Date'})
    pivot['_s'] = pd.to_datetime(pivot['Date'], format='%m/%d/%y', errors='coerce')
    pivot = pivot.sort_values('_s').drop(columns=['_s'])
    pivot['Total'] = pivot[DISPLAY_COLS].sum(axis=1)
    total_row = {'Date': 'Total'}
    for c in DISPLAY_COLS + ['Total']:
        total_row[c] = int(pivot[c].sum()) if c != 'Date' else 'Total'
    monthly_df = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)
    monthly_cols = ['Date'] + DISPLAY_COLS + ['Total']
    monthly_records = df_to_records(monthly_df[monthly_cols])
    ytd = nr.groupby(['month', 'display_col']).size().reset_index(name='count')
    ytd['month_str'] = ytd['month'].astype(str)
    ytd_pivot = ytd.pivot_table(index='display_col', columns='month_str', values='count', fill_value=0)
    ytd_pivot = ytd_pivot.reindex(DISPLAY_COLS, fill_value=0)
    month_str_list = [str(m) for m in all_months]
    grand_totals = {m: float(ytd_pivot[m].sum()) if m in ytd_pivot.columns else 0.0 for m in month_str_list}
    ytd_pct_rows = []
    for col in PCT_COLS:
        row = {'Reason': '%' + col}
        for m in month_str_list:
            if m in ytd_pivot.columns and grand_totals.get(m, 0) > 0:
                val = float(ytd_pivot.loc[col, m]) if col in ytd_pivot.index else 0.0
                row[m] = round(val / grand_totals[m] * 100, 2)
            else:
                row[m] = 0.0
        ytd_pct_rows.append(row)
    ytd_cols = ['Reason'] + month_str_list
    chart_datasets = [{'label': row['Reason'], 'data': [row.get(m, 0.0) for m in month_str_list]} for row in ytd_pct_rows]
    return {
        'report': '1_6', 'title': f'1.6 {agent_name} Not Ready Count',
        'agent_name': agent_name, 'report_month': str(report_month),
        'monthly_columns': monthly_cols, 'monthly_table': monthly_records,
        'ytd_columns': ytd_cols, 'ytd_table': ytd_pct_rows,
        'chart_labels': month_str_list, 'chart_datasets': chart_datasets,
        'table': monthly_records, 'columns': monthly_cols,
    }


def _empty_result():
    return {
        'report': '1_6', 'title': '1.6 Not Ready Count', 'agent_name': 'Agent',
        'report_month': '', 'monthly_columns': [], 'monthly_table': [],
        'ytd_columns': [], 'ytd_table': [], 'chart_labels': [], 'chart_datasets': [],
        'table': [], 'columns': [],
    }


def _fix_chart_axes_xml(xlsx_bytes):
    """Fix openpyxl chart XML for Excel using ElementTree.
    - catAx: axPos l->b, add delete+crosses in correct OOXML element order
    - valAx: add delete+crosses in correct OOXML element order
    Element ordering matters: Excel is strict about OOXML CT_CatAx/CT_ValAx sequence.
    """
    import zipfile as _zf
    import xml.etree.ElementTree as _ET
    import re as _re

    _NS = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
    _CATAX_ORDER = [
        'axId','scaling','delete','axPos',
        'majorGridlines','minorGridlines','title','numFmt',
        'majorTickMark','minorTickMark','tickLblPos','spPr','txPr',
        'crossAx','crosses','crossesAt','auto','lblAlgn',
        'lblOffset','tickLblSkip','tickMarkSkip','noMultiLvlLbl','extLst'
    ]
    _VALAX_ORDER = [
        'axId','scaling','delete','axPos',
        'majorGridlines','minorGridlines','title','numFmt',
        'majorTickMark','minorTickMark','tickLblPos','spPr','txPr',
        'crossAx','crossBetween','crosses','crossesAt',
        'majorUnit','minorUnit','dispUnits','extLst'
    ]

    def _fix_ax(ax_elem, order, axpos_val):
        tag_map = {}
        for child in list(ax_elem):
            local = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            tag_map[local] = child
        if 'axPos' in tag_map:
            tag_map['axPos'].set('val', axpos_val)
        else:
            el = _ET.Element(f'{{{_NS}}}axPos')
            el.set('val', axpos_val)
            tag_map['axPos'] = el
        if 'delete' not in tag_map:
            el = _ET.Element(f'{{{_NS}}}delete')
            el.set('val', '0')
            tag_map['delete'] = el
        if 'crosses' not in tag_map and 'crossesAt' not in tag_map:
            el = _ET.Element(f'{{{_NS}}}crosses')
            el.set('val', 'autoZero')
            tag_map['crosses'] = el
        for child in list(ax_elem):
            ax_elem.remove(child)
        for key in order:
            if key in tag_map:
                ax_elem.append(tag_map[key])

    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with _zf.ZipFile(buf, 'r') as zin, _zf.ZipFile(out, 'w', _zf.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                xml_str = raw.decode('utf-8')
                # Register all namespaces from original to preserve prefixes
                for prefix, uri in _re.findall(r'xmlns(?::(\w+))?="([^"]+)"', xml_str):
                    _ET.register_namespace(prefix, uri)
                root = _ET.fromstring(xml_str)
                for catax in root.iter(f'{{{_NS}}}catAx'):
                    _fix_ax(catax, _CATAX_ORDER, 'b')
                for valax in root.iter(f'{{{_NS}}}valAx'):
                    _fix_ax(valax, _VALAX_ORDER, 'l')
                raw = (b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\n"
                       + _ET.tostring(root, encoding='unicode').encode('utf-8'))
            zout.writestr(item, raw)
    return out.getvalue()


def export_excel(payload):
    data = payload.get('data', payload)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel

    BLUE = '1F4E79'; LIGHT_BLUE = 'BDD7EE'; WHITE = 'FFFFFF'; LGRAY = 'F2F2F2'
    CHART_COLORS = ['4472C4', 'ED7D31', 'A9D18E', 'FF0000', '7030A0']
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    def write_table(ws, title_text, columns, records, has_total_row=False):
        ws.cell(1, 1, title_text).font = Font(bold=True, size=13, color=BLUE)
        if len(columns) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.cell(1, 1).alignment = Alignment(horizontal='center')
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(2, ci, col)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = PatternFill('solid', fgColor=BLUE)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin
        ws.row_dimensions[2].height = 30
        n_rec = len(records)
        for ri, rec in enumerate(records, start=3):
            is_total = has_total_row and (ri == n_rec + 2)
            for ci, col in enumerate(columns, 1):
                val = rec.get(col)
                cell = ws.cell(ri, ci, val)
                cell.border = thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if is_total:
                    cell.font = Font(bold=True, color=BLUE)
                    cell.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
                elif ri % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor=LGRAY)
        for ci, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col)) + 4, 14)

    wb = Workbook()
    agent = data.get('agent_name', 'Agent')
    month = data.get('report_month', '')
    monthly_columns = data.get('monthly_columns', [])
    monthly_table = data.get('monthly_table', [])
    ytd_columns = data.get('ytd_columns', [])
    ytd_table = data.get('ytd_table', [])

    # Sheet 1: Monthly
    ws1 = wb.active
    ws1.title = 'Monthly'
    write_table(ws1, f'{agent} Not Ready Count: {month}', monthly_columns, monthly_table, has_total_row=True)

    n_data = len(monthly_table)
    n_data_only = n_data - 1  # exclude Total row

    if n_data_only > 0 and len(monthly_columns) > 1:
        display_col_indices = [i + 1 for i, c in enumerate(monthly_columns) if c in DISPLAY_COLS]
        if display_col_indices:
            chart1 = BarChart()
            chart1.type = 'col'
            chart1.grouping = 'stacked'
            chart1.overlap = 100
            chart1.title = f'{agent} Not Ready Count: {month}'
            chart1.y_axis.title = 'Count'
            chart1.x_axis.title = 'Date'
            chart1.width = 22
            chart1.height = 14
            for col_idx in display_col_indices:
                col_name = monthly_columns[col_idx - 1]
                ref = Reference(ws1, min_col=col_idx, max_col=col_idx,
                                min_row=3, max_row=2 + n_data_only)
                chart1.add_data(ref, titles_from_data=False)
                chart1.series[-1].tx = SeriesLabel(v=col_name)
            chart1.set_categories(Reference(ws1, min_col=1, min_row=3, max_row=2 + n_data_only))
            for i, s in enumerate(chart1.series):
                color = CHART_COLORS[i % len(CHART_COLORS)]
                s.graphicalProperties.solidFill = color
                s.graphicalProperties.line.solidFill = color
            ws1.add_chart(chart1, f'A{n_data + 5}')

    # Sheet 2: YTD %
    ws2 = wb.create_sheet('YTD %')
    write_table(ws2, f'{agent} Not Ready Reasons YTD %', ytd_columns, ytd_table, has_total_row=False)

    n_ytd = len(ytd_table)
    n_months = len(ytd_columns) - 1

    if n_ytd > 0 and n_months > 0:
        chart2 = BarChart()
        chart2.type = 'col'
        chart2.grouping = 'clustered'
        chart2.title = f'{agent} : Not Ready Reasons YTD (%)'
        chart2.y_axis.title = 'Percent (%)'
        chart2.x_axis.title = 'Month'
        chart2.width = 22
        chart2.height = 14
        for row_idx in range(3, 3 + n_ytd):
            ref = Reference(ws2, min_col=2, max_col=len(ytd_columns),
                            min_row=row_idx, max_row=row_idx)
            chart2.add_data(ref, titles_from_data=False, from_rows=True)
        chart2.set_categories(Reference(ws2, min_col=2, max_col=len(ytd_columns), min_row=2, max_row=2))
        for i, s in enumerate(chart2.series):
            label = ytd_table[i].get('Reason', f'Row {i+1}') if i < len(ytd_table) else f'Row {i+1}'
            s.tx = SeriesLabel(v=label)
            color = CHART_COLORS[i % len(CHART_COLORS)]
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.solidFill = color
        ws2.add_chart(chart2, f'A{n_ytd + 5}')

    out = io.BytesIO()
    wb.save(out)
    return _fix_chart_axes_xml(out.getvalue())
